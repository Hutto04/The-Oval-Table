from openpyxl import load_workbook
from numpy import *
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split #10-01-23
from sklearn.metrics import mean_squared_error #10-01-23
from sklearn.metrics import r2_score #10-03-2023

wb = load_workbook(filename = 'cleaned_dataset.xlsx')

sheet = wb.active

#print(sheet.cell(2,2,).value)

counter = 1

for x in range(1, sheet.max_row):
        counter = counter + 1
       

#print(counter)

lst = []

lst.append([1,2,3])
lst.append([4,5,6])
#print(lst[1])

mainlst = []
templst= []
#print("row count & column count:")
#print(sheet.max_row)
#print(sheet.max_column)

for column in range(1, sheet.max_column+1):
    for row in range(1, sheet.max_row+1):
        cellData = sheet.cell(row, column).value
        templst.append(cellData)
    mainlst.append(templst)
    templst = []

#print(mainlst[sheet.max_column-1][sheet.max_row-1])

selectfeature = 95
firstfeature = mainlst[selectfeature][1:] #x rpi
firstfeature = np.array(firstfeature)
response = mainlst[-1][1:] #y oil
#plt.figure(figsize=(1,1))
#plt.scatter(firstfeature, response)
#plt.show()

poly = PolynomialFeatures(degree=2, include_bias=False)

poly_features = poly.fit_transform(firstfeature.reshape(-1, 1))
poly_reg_model = LinearRegression()
poly_reg_model.fit(poly_features, response)

y_predicted = poly_reg_model.predict(poly_features)

xtitle = mainlst[selectfeature][:1]
plt.figure(figsize=(10,6))
plt.title("Feature vs. Real Oil Prices")
plt.xlabel(xtitle)
plt.ylabel("Real Oil Prices")
plt.scatter(firstfeature, response)
plt.plot(firstfeature, y_predicted, c="red")
plt.show()

#10-01-2023 new stuff
secfeat = mainlst[2][1:]
secfeat = np.array(secfeat)
response = np.array(response)

df = pd.DataFrame({"first_feature":firstfeature.reshape(142,), "second_feature":secfeat.reshape(142,), "response":response.reshape(142,)}, index=range(0,142))

X, y = df[["first_feature", "second_feature"]], df["response"]

poly = PolynomialFeatures(degree=2, include_bias=False)
poly_features = poly.fit_transform(X)
X_train, X_test, y_train, y_test = train_test_split(poly_features, y, test_size=0.3, random_state=42)
poly_reg_model = LinearRegression()
poly_reg_model.fit(X_train, y_train)

poly_reg_y_predicted = poly_reg_model.predict(X_test)
poly_reg_rmse = np.sqrt(mean_squared_error(y_test, poly_reg_y_predicted))
print(poly_reg_rmse)
